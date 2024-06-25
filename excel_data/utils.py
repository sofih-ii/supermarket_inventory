import os
from openpyxl import load_workbook

def update_master_excel(inventory_file, master_file):

    wb_master = load_workbook(filename=master_file)
    sheet_master = wb_master.active

    wb_inventory = load_workbook(filename=inventory_file)
    sheet_inventory = wb_inventory.active

    for row in sheet_inventory.iter_rows(min_row=2, values_only=True):
        sheet_master.append(row)

    wb_master.save(master_file)